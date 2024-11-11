from openpyxl import load_workbook
import psycopg2 as pg
import pandas as pd
from sqlalchemy import create_engine
from typing import List, Dict, Union, Literal
from datetime import datetime

class ReadExcelSheetException(Exception):
    pass

class OrientationError(Exception):
    pass

class MatrixStructureException(Exception):
    pass

class DifferentLengthsExceptions(Exception):
    pass

class XlsxPgLoader:
    def __init__(self, file: str,  __host: str, __dbname: str, __user: str, __password: str, __port: int):
        self.file = file
        self.connection = pg.connect(host=__host, dbname=__dbname, user=__user, password=__password, port=__port)
        self.cursor = self.connection.cursor()
        self.engine = create_engine(f'postgresql://{__user}:{__password}@{__host}:{__port}/{__dbname}')
        self.workbook = load_workbook(self.file, data_only=True)

    def getSheetList(self) -> list:
        return self.workbook.sheetnames

    def checkRangeType(self, *args) -> str:
        try:
            args = list(args)

            if len(args) == 1:
                type = 'field'

            elif len(args) == 2:
                dim1, dim2 = args[0], args[1]

                if dim1 == dim2:
                    type = 'field'
                else:
                    for i in range(len(dim1)):
                        if dim1[i].isalpha() == False:
                            dim1alpha = dim1[: i]
                            dim1num = dim1[i: ]
                            break

                    for i in range(len(dim2)):
                        if dim2[i].isalpha() == False:
                            dim2alpha = dim2[: i]
                            dim2num = dim2[i: ]
                            break

                    if dim1alpha == dim2alpha:
                        type = 'col'
                    elif dim1num == dim2num:
                        type = 'row'
                    else:
                        type = 'tab'

            else:
                raise AttributeError(f'Minimum number of arguments is 1, maximum - 2, Given {len(args)}')
            return type
        except AttributeError as e:
            raise e

    def getColumnNamesOrTypes(self, sheetname: str, colNameStart: str = None, colNameEnd: str = None) -> List[str]:
        try:
            sheet = self.workbook[sheetname]
            data = []

            if not colNameStart and not colNameEnd:
                raise MatrixStructureException("Expected a row or column range, not the entire sheet.")
            elif colNameStart and not colNameEnd:
                data.append([str(sheet[colNameStart].value).lower()])
            elif not colNameStart and colNameEnd:
                data.append([str(sheet[colNameEnd].value).lower()])
            else:
                try:
                    typeRange = self.checkRangeType(colNameStart, colNameEnd)

                    if typeRange == 'row':
                        for row in sheet[colNameStart: colNameEnd]:
                            data.append([str(cell.value).lower() for cell in row])
                    elif typeRange == 'col':
                        for col in zip(*[row for row in sheet[colNameStart: colNameEnd]]):
                            data.append([str(cell.value).lower() for cell in col])
                    elif typeRange == 'field':
                        data.append([str(sheet[colNameStart].value).lower()])
                    else:
                        raise MatrixStructureException(
                            f"Range '{colNameStart}:{colNameEnd}' must be a row or column, not a matrix.")
                except UnboundLocalError:
                    raise MatrixStructureException(
                        "Coordinate format error. Expected format: letter for column and number for row (e.g., 'B5').")

            data = data[0]
            return data

        except KeyError:
            raise ReadExcelSheetException(f"Sheet '{sheetname}' in file '{self.file}' does not exist.")

    def getDataFromSheetToDataFrame(self, orientation: Literal['vertical', 'horizontal'], sheetname: str,
                                    colNameStart: str = None, colNameEnd: str = None,
                                    dataStart: str = None, dataEnd: str = None) -> dict:
        try:
            sheet = self.workbook[sheetname]
            columnNames = self.getColumnNamesOrTypes(sheetname, colNameStart=colNameStart, colNameEnd=colNameEnd)

            data = {}
            i = 0

            if not (dataStart or dataEnd):
                raise MatrixStructureException("There must be min 1 argument set - dataStart or/and dataEnd")

            match orientation:
                case 'horizontal':
                    for row in sheet[dataStart: dataEnd]:
                        key = columnNames[i] if i < len(columnNames) else f'column{i + 1}'
                        data[key] = [cell.value for cell in row]
                        i += 1
                case 'vertical':
                    for col in zip(*[row for row in sheet[dataStart: dataEnd]]):
                        key = columnNames[i] if i < len(columnNames) else f'column{i + 1}'
                        data[key] = [cell.value for cell in col]
                        i += 1
                case _:
                    raise OrientationError(f"Orientation must be 'vertical' or 'horizontal', not '{orientation}'")

            if len(data) != len(columnNames):
                raise DifferentLengthsExceptions(
                    f"Length of column list ({len(columnNames)}) is not equal to length of dictionary ({len(data)})"
                )

            return data

        except UnboundLocalError:
            print('UnboundLocalError: Wrong format of coordinates - the first value is column name as a letter, '
                  'and the second - row number, for example B5, A10, etc.')
        except KeyError:
            raise ReadExcelSheetException
        except TypeError:
            raise MatrixStructureException

    def buildDataFrame(self, data: dict) -> pd.DataFrame:
        return pd.DataFrame(data)

    def createTableSql(self, sheetname: str, columnsRange: tuple, typesRange: tuple,
                       table: str = 'table_name', schema: str = 'public',
                       constraints: Dict[str, Union[List[str], Dict[str, Dict[str, str]], str]] = None,
                       isPartitioned: bool = False, partitionType: str = None,
                       partitionColumns: List[str] = None, partitions: List[Dict[str, str]] = None,
                       save: bool = False) -> Union[str, List[str]]:
        try:
            columnNames = self.getColumnNamesOrTypes(sheetname, *columnsRange)
            dataTypes = self.getColumnNamesOrTypes(sheetname, *typesRange)

            if len(columnNames) != len(dataTypes):
                raise DifferentLengthsExceptions("The number of columns is not equal to the number of its types")

            columnsSql = [f"{colName} {dataType}" for colName, dataType in zip(columnNames, dataTypes)]

            if constraints:
                for constraintType, value in constraints.items():
                    if constraintType.lower().startswith('p') and isinstance(value, list):
                        columnsSql.append(f"PRIMARY KEY ({', '.join(value)})")

                    elif constraintType.lower().startswith('f') and isinstance(value, dict):
                        for fkName, fkDetails in value.items():
                            referencedTable = fkDetails.get("table")
                            referencedColumn = fkDetails.get("column")
                            if referencedTable and referencedColumn:
                                columnsSql.append(
                                    f"CONSTRAINT {fkName} FOREIGN KEY ({', '.join(fkDetails['columns'])}) "
                                    f'REFERENCES {referencedTable} ({referencedColumn})'
                                )

                    elif constraintType.lower().startswith('c') and isinstance(value, str):
                        columnsSql.append(f"CHECK ({value})")

            partitionSql = ""
            if isPartitioned and partitionType and partitionColumns:
                partitionSql = f" PARTITION BY {partitionType} ({', '.join(partitionColumns)})"
                partitionDefinitions = []

                for part in partitions or []:
                    partition_name = part.get("name")
                    values = part.get("values")
                    if partition_name and values:
                        if partitionType.lower() == "range":
                            partitionDefinitions.append(f"PARTITION {partition_name} VALUES LESS THAN ({values})")
                        elif partitionType.lower() == "list":
                            partitionDefinitions.append(f"PARTITION {partition_name} VALUES IN ({values})")

                if partitionDefinitions:
                    partitionSql += " (\n  " + ",\n  ".join(partitionDefinitions) + "\n)"


            createTableSql = f'DROP TABLE IF EXISTS {schema}.{table};\n'
            createTableSql += f'CREATE TABLE {schema}.{table} (\n  ' + ',\n  '.join(columnsSql) + f'\n){partitionSql};'

            if save:
                scriptName = f"script_postgres_{table.replace('.', '_')}.sql"
                try:
                    with open(scriptName, 'w') as file:
                        file.truncate()
                except FileExistsError:
                    print('File not exists')

                with open(scriptName, 'a', encoding='UTF8') as file:
                    file.write(f'{createTableSql}\n\n')

                return [createTableSql, scriptName]

            else:
                return createTableSql

        except DifferentLengthsExceptions as e:
            print(f"Błąd: {e}")
            return ""
        except Exception as e:
            print(f"Wystąpił nieoczekiwany błąd: {e}")
            return ""

    def insertData(self, df: pd.DataFrame, table: str, schema: str = 'public', ifExists:
                Literal['fail', 'replace', 'append'] = 'append') -> None:

        try:
            today = datetime.now()
            print(f'Start: {today}\n{table.upper()} LOADING...\n')
            df.to_sql(con=self.engine, schema=schema, name=table, if_exists=ifExists, index=False)
            today = datetime.now()
            print(f'ETL Successed\nEnd: {today}\nInserted: {df.count()} rows')
        except Exception as e:
            print(f'Error\nEnd: {today}\nInserted 0 rows\n')
            raise e

    def __del__(self):
        self.cursor.close()
        self.connection.close()
        del self.engine